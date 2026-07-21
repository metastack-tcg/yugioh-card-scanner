"""Seller-facing buylist workbook: price every candidate printing of each
scanned card (via the TCGP Scraper package), then write an .xlsx where the
seller picks the printing from a dropdown and Set / Price / Verified / Proof
update by formula. No app needed on their end — just Excel.

    python buylist.py --selftest   # offline check of the workbook writer
"""
import json
import os
import re
import string
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

# dev runs import the scraper from its sibling repo; the frozen exe has the
# package bundled by PyInstaller (--paths), so the sibling won't exist there
_SCRAPER = Path(__file__).parent.parent / "TCGP Scraper"
if _SCRAPER.exists():
    sys.path.insert(0, str(_SCRAPER))
from ygo_tcgplayer_pricer import tcgplayer_catalog, tcgplayer_pricing  # noqa: E402

CONDITIONS = tcgplayer_pricing.CONDITIONS  # NM..Damaged — seller picks per row in Excel
DEFAULT_CONDITION = tcgplayer_pricing.DEFAULT_CONDITION

INDEX_CACHE = Path(os.environ.get("APPDATA", Path.home())) / "metastack card scanner" / "tcgp_index_v2.json"
INDEX_MAX_AGE = 7 * 86400  # ponytail: weekly rebuild; new sets mostly match by group name anyway
_IDX = None


def _global_index():
    """(set_code, rarity) -> product across the ENTIRE TCGPlayer Yu-Gi-Oh
    catalog. Fallback for printings whose set can't be matched by group name —
    e.g. video-game promos TCGPlayer files under catch-all groups. Set codes
    are unique game-wide, so this is safe. ~600 tcgcsv fetches on first build,
    then cached to disk."""
    global _IDX
    if _IDX is not None:
        return _IDX
    if INDEX_CACHE.exists() and time.time() - INDEX_CACHE.stat().st_mtime < INDEX_MAX_AGE:
        try:
            raw = json.loads(INDEX_CACHE.read_text())
            _IDX = tuple({tuple(k.split("|", 1)): v for k, v in part.items()}
                         for part in (raw["base"], raw["ea"]))
            return _IDX
        except Exception:
            pass  # corrupt cache — rebuild below

    base, ea = {}, {}
    gname = {g["groupId"]: g["name"] for g in tcgplayer_catalog._all_groups()}
    failed = 0
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(tcgplayer_catalog.get_product_lookup, gid): gid
                   for gid in gname}
        for f in as_completed(futures):
            gid = futures[f]
            try:
                by_r, ea_r, _ = f.result()
            except Exception:
                failed += 1  # one throttled group must not sink the other 650
                continue
            for k, v in by_r.items():
                base.setdefault(k, dict(v, group=gname[gid]))
            for k, v in ea_r.items():
                ea.setdefault(k, dict(v, group=gname[gid]))
    if failed and failed >= len(gname) // 2:
        raise RuntimeError(f"TCGPlayer catalog unreachable ({failed} groups failed)")
    INDEX_CACHE.parent.mkdir(parents=True, exist_ok=True)
    INDEX_CACHE.write_text(json.dumps(
        {"base": {"|".join(k): v for k, v in base.items()},
         "ea": {"|".join(k): v for k, v in ea.items()}}))
    _IDX = (base, ea)
    return _IDX


def _price_jobs(jobs, progress=None):
    """{(product_id, condition): priced_or_None} — concurrent, deduped by caller."""
    prices = {}
    with ThreadPoolExecutor(max_workers=tcgplayer_pricing.MAX_WORKERS) as pool:
        futures = {pool.submit(tcgplayer_pricing.get_lowest_price_safe,
                               pid, condition=cond): (pid, cond)
                   for pid, cond in jobs}
        for i, f in enumerate(as_completed(futures), 1):
            prices[futures[f]] = f.result()
            if progress:
                progress(i, len(jobs))
    return prices


def tcgp_candidates(name, read_code, min_ratio=0.75):
    """Candidate printings straight from the TCGPlayer catalog, keyed by the
    set code read off the physical card. Fallback for when ygoprodeck's
    per-card set list is missing the printing in hand (their lists lag —
    e.g. Dark Magician's omits Maximum Gold: El Dorado entirely). The code
    pattern narrows the field; a name-similarity floor keeps out neighbours
    on the same page."""
    import difflib

    import scan_cards

    gbase, _ = _global_index()
    out = {}
    for (number, rarity), prod in gbase.items():
        # an empty read_code matches every number — name-only mode, for cards
        # ygoprodeck knows but has no printings for (brand-new sets)
        if read_code and not scan_cards.code_matches(read_code, number):
            continue
        pname = prod.get("name", "").split(" (")[0]  # strip "(Extended Art)" etc.
        m = difflib.SequenceMatcher(None, name.lower(), pname.lower())
        if m.quick_ratio() < min_ratio or m.ratio() < min_ratio:  # quick_ratio gates the slow call
            continue
        e = out.setdefault(number, {"set_code": number,
                                    "set_name": prod.get("group", ""), "rarities": []})
        pretty = string.capwords(rarity)  # not .title(): "Collector's", not "Collector'S"
        if pretty not in e["rarities"]:
            e["rarities"].append(pretty)
        # carry the best-matching product name so a misread card name can be
        # corrected from the TCGPlayer catalog when ygoprodeck drew a blank
        if m.ratio() > e.get("_score", 0):
            e["_score"], e["card_name"] = m.ratio(), pname
    return [dict({k: v for k, v in e.items() if k != "_score"},
                 rarities=sorted(e["rarities"])) for _, e in sorted(out.items())]


def resolve_candidates(name, code, snippet=None):
    """The full lookup chain, shared by scan and refresh: ygoprodeck first,
    then the TCGPlayer catalog — as a REPLACEMENT when the read code matches
    none of ygoprodeck's printings, and as a MERGE when there was no code at
    all, because ygoprodeck's per-card set lists lag (Dododo Warrior's omits
    the Duelist's Advance reprint entirely). The merge threshold is 0.9:
    tight enough to keep 'Dark Magician Girl' out of Dark Magician's list,
    loose enough to survive TCGPlayer's own typos ('Dodododo Warrior')."""
    import scan_cards

    try:
        cname, cands = scan_cards.db_lookup(name, code, snippet)
    except Exception:
        cname, cands = name, []
    def merged(cands):
        known = {k["set_code"] for k in cands}
        return cands + [k for k in tcgp_candidates(cname, "", min_ratio=0.9)
                        if k["set_code"] not in known]

    try:
        if not cands or (code and not any(
                scan_cards.code_matches(code, k["set_code"]) for k in cands)):
            extra = tcgp_candidates(cname, code or "")
            if extra:
                if not cands and extra[0].get("card_name"):
                    cname = extra[0]["card_name"]  # trust the catalog's spelling
                cands = extra
            else:
                # the code matches nothing anywhere — a misread; treat it as
                # absent so every known printing of the name is still on offer
                cands = merged(cands)
        elif not code:
            cands = merged(cands)
    except Exception:
        pass  # a catalog hiccup must not lose the ygoprodeck result
    return cname, cands


def build_options(cards, progress=None):
    """Attach card["options"] — one entry per candidate (printing, rarity),
    priced. Options span every candidate set so the seller can still correct
    a wrong auto-resolution in Excel."""
    for card in cards:
        card["options"] = []
        for cand in card["sets"]:
            gid = tcgplayer_catalog.find_group_id(cand["set_name"])
            by_rarity, ea_by_rarity, _ = (
                tcgplayer_catalog.get_product_lookup(gid) if gid else ({}, {}, {}))
            for rarity in cand["rarities"]:
                key = (cand["set_code"], rarity.strip().lower())
                base = by_rarity.get(key)
                ea_key = ea_by_rarity.get(key)
                if base is None and ea_key is None:
                    gbase, gea = _global_index()
                    base, ea_key = gbase.get(key), gea.get(key)
                label = f"{cand['set_code']} · {rarity}"
                if base and base.get("is_extended_art"):
                    label += " (Extended Art)"
                card["options"].append(
                    {"label": label, "set_name": cand["set_name"], "product": base})
                if ea_key:
                    card["options"].append(
                        {"label": f"{cand['set_code']} · {rarity} (Extended Art)",
                         "set_name": cand["set_name"], "product": ea_key})

    # one request per (product, condition) — 5x the calls of NM-only, so a
    # binder page lands around 100-250 requests; dedup + the scraper's thread
    # pool keep it to about a minute
    jobs = {(o["product"]["productId"], cond)
            for c in cards for o in c["options"] if o["product"]
            for cond in CONDITIONS}
    prices = _price_jobs(jobs, progress)

    for c in cards:
        for o in c["options"]:
            o["url"] = o["product"]["url"] if o["product"] else None
            o["prices"], o["verified"] = {}, {}
            for cond in CONDITIONS:
                priced = prices.get((o["product"]["productId"], cond)) if o["product"] else None
                o["prices"][cond] = priced["price"] if priced else None
                o["verified"][cond] = priced["verified"] if priced else None


HEADERS = ["Card Name", "Printing (choose)", "Set", "Condition", "Qty",
           "Price/Unit", "Price", "Verified", "Link/Proof", "Photo"]

# The $5-per-thousand bulk tier: paper rarities plus the bulk-binder foil
# patterns. Every other rarity of a sub-pricepoint card lands at $30/1,000.
# Cards at/above the pricepoint never touch this list.
BULK_CR = ["Common", "Rare", "Short Print", "Super Short Print",
           "Normal Parallel Rare", "Duel Terminal Normal Parallel Rare",
           "Duel Terminal Rare Parallel Rare", "Duel Terminal Technology Common",
           "Starfoil Rare", "Shatterfoil Rare", "Mosaic Rare", "Parallel Rare"]
TINT = "F6ECE4"  # house accent tint — marks rows still needing a choice


def _option_row(o):
    """One Options-sheet row for an option dict from build_options."""
    def vtext(cond):
        if o["prices"][cond] is None:
            return ""
        return "yes" if o["verified"][cond] else "no"
    return ([o["label"], o["set_name"], o["url"] or ""]
            + [o["prices"][cond] for cond in CONDITIONS]
            + [vtext(cond) for cond in CONDITIONS])


def _wire_row(ws, r, lo, hi):
    """Formulas + printing dropdown for a Cards row whose options occupy
    Options rows lo..hi. Shared by export and refresh so they can't drift.
    The IF(...="","",...) wraps exist because INDEX/VLOOKUP render an empty
    cell as 0, and a missing listing must read as blank, not $0.00."""
    from openpyxl.worksheet.datavalidation import DataValidation

    labels = f"Options!$A${lo}:$A${hi}"
    row_m = f"MATCH($B{r},{labels},0)"
    col_m = f"MATCH($D{r},Options!$D$1:$H$1,0)"
    price_ix = f"INDEX(Options!$D${lo}:$H${hi},{row_m},{col_m})"
    verif_ix = f"INDEX(Options!$I${lo}:$M${hi},{row_m},{col_m})"
    url_vl = f"VLOOKUP($B{r},Options!$A${lo}:$C${hi},3,FALSE)"
    ws.cell(r, 3).value = (f'=IFERROR(VLOOKUP($B{r},Options!$A${lo}:$B${hi},2,FALSE),"")')
    ws.cell(r, 6).value = f'=IFERROR(IF({price_ix}="","",{price_ix}),"")'
    ws.cell(r, 6).number_format = "$0.00"
    ws.cell(r, 7).value = f'=IFERROR($E{r}*$F{r},"")'
    ws.cell(r, 7).number_format = "$0.00"
    ws.cell(r, 8).value = f'=IFERROR(IF({verif_ix}="","",{verif_ix}),"")'
    ws.cell(r, 9).value = f'=IFERROR(IF({url_vl}="","",HYPERLINK({url_vl})),"")'
    # hidden class column for the bulk split: rarity is the text after the
    # "·" in the chosen printing, matched against BULK_CR
    rar = f'TRIM(MID($B{r},FIND("·",$B{r})+1,99))'
    tier = "{" + ",".join(f'"{x}"' for x in BULK_CR) + "}"
    ws.cell(r, 11).value = (
        f'=IF($B{r}="","",IF(ISNUMBER(MATCH({rar},{tier},0)),"cr","foil"))')
    dv = DataValidation(type="list", formula1=labels,
                        allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.add(ws.cell(r, 2).coordinate)


def write_workbook(path, cards, timestamp=""):
    """cards must already carry ["options"] from build_options."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Cards"
    opts = wb.create_sheet("Options")
    opts.sheet_state = "hidden"
    # A label · B set · C url · D-H price per condition · I-M verified per condition
    opts.append(["Label", "Set", "URL"] + CONDITIONS + CONDITIONS)

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if timestamp:
        ws.append([])  # spacer under a note row keeps the table clean
        ws["A2"] = ("Prices: lowest verified TCGPlayer listing in the chosen "
                    f"condition, {timestamp}. A blank price means no verified "
                    "listing right now.")
        ws["A2"].font = Font(italic=True, size=9)

    cond_dv = DataValidation(
        type="list", formula1='"' + ",".join(CONDITIONS) + '"',
        allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(cond_dv)

    r = 3 if timestamp else 2
    orow = 2
    for c in cards:
        ws.cell(r, 1, c["name"])
        ws.cell(r, 5, c.get("qty", 1))
        ws.cell(r, 10, c["photo"])
        ws.cell(r, 12, c["name"])  # hidden original name — edit detection on refresh
        if c["options"]:
            lo, hi = orow, orow + len(c["options"]) - 1
            for o in c["options"]:
                opts.append(_option_row(o))
            orow = hi + 1
            _wire_row(ws, r, lo, hi)
            cond_dv.add(ws.cell(r, 4).coordinate)
            ws.cell(r, 4).value = DEFAULT_CONDITION

            labels = [o["label"] for o in c["options"]]
            pre = f"{c['set_code']} · {c['rarity']}" if c["set_code"] and c["rarity"] else None
            # an Extended Art foil read outranks the plain printing of the same rarity
            tries = (pre, pre and pre + " (Extended Art)")
            if pre and "extended art" in (c.get("guess") or "").lower():
                tries = (pre + " (Extended Art)", pre)
            pick = next((v for v in tries if v in labels), None)
            if pick:
                ws.cell(r, 2).value = pick
            else:
                ws.cell(r, 2).fill = PatternFill("solid", fgColor=TINT)
        r += 1

    first = 3 if timestamp else 2
    last = r - 1
    if r > first:
        ws.cell(r + 1, 1, "Total").font = Font(bold=True)
        ws.cell(r + 1, 2).value = (
            f'=COUNTA(B{first}:B{last})&" of "&COUNTA(A{first}:A{last})&" cards chosen"')
        ws.cell(r + 1, 5).value = f"=SUM(E{first}:E{last})"  # total physical cards
        total = ws.cell(r + 1, 7)
        total.value = f"=SUM(G{first}:G{last})"
        total.number_format = "$0.00"
        total.font = Font(bold=True)

        # --- the offer: bulk under a pricepoint, a percentage above it ------
        tint = PatternFill("solid", fgColor=TINT)
        pp, rt = r + 3, r + 4
        num = f"ISNUMBER(F{first}:F{last})"
        under = f"{num}*(F{first}:F{last}<$F${pp})"
        over = f"{num}*(F{first}:F{last}>=$F${pp})"
        cls = f"$K${first}:$K${last}"
        qty, line = f"E{first}:E{last}", f"G{first}:G{last}"

        def label(row, text, bold=False):
            cell = ws.cell(row, 1, text)
            if bold:
                cell.font = Font(bold=True)

        label(pp, "Bulk pricepoint (cards under this are bulk)")
        c = ws.cell(pp, 6)
        c.value, c.number_format, c.fill = 1.0, "$0.00", tint
        label(rt, "Offer rate for cards at/above the pricepoint")
        c = ws.cell(rt, 6)
        c.value, c.number_format, c.fill = 0.65, "0%", tint

        rows = [
            (r + 5, "Bulk commons & rares — $5 per 1,000",
             f'=SUMPRODUCT({under}*({cls}="cr"),{qty})', f"=E{r + 5}*5/1000"),
            (r + 6, "Bulk foils & others — $30 per 1,000",
             f'=SUMPRODUCT({under}*({cls}="foil"),{qty})', f"=E{r + 6}*30/1000"),
            (r + 7, "Cards at/above the pricepoint (market value)",
             f"=SUMPRODUCT({over},{qty})", f"=SUMPRODUCT({over},{line})"),
        ]
        for row, text, qf, gf in rows:
            label(row, text)
            ws.cell(row, 5).value = qf
            g = ws.cell(row, 7)
            g.value, g.number_format = gf, "$0.00"
        label(r + 8, "OFFER TOTAL", bold=True)
        g = ws.cell(r + 8, 7)
        g.value = f"=G{r + 5}+G{r + 6}+G{r + 7}*$F${rt}"
        g.number_format, g.font = "$0.00", Font(bold=True)

    for col, width in zip("ABCDEFGHIJ", (32, 34, 34, 13, 5, 10, 10, 9, 40, 16)):
        ws.column_dimensions[col].width = width
    ws.column_dimensions["K"].hidden = True
    ws.column_dimensions["L"].hidden = True
    wb.save(path)


PRODUCT_URL_RE = re.compile(r"/product/(\d+)")
SET_CODE_RE = re.compile(r"\b[A-Z0-9]{2,5}-[A-Z0-9?]{2,8}\b", re.I)
OPT_RANGE_RE = re.compile(r"Options!\$A\$(\d+):\$B\$(\d+)")


def refresh_workbook(path, timestamp="", progress=None):
    """Re-price a workbook the seller sent back — and honor their corrections.

    A row is RE-RESOLVED (fresh candidates, new dropdown, new prices) when the
    seller edited the card name (vs the hidden original in column L), typed a
    printing that isn't one of the row's options, or the row never had options.
    Everything else keeps its choices and simply gets fresh prices.

    All dropdowns are rebuilt from scratch: Excel re-saves cross-sheet
    validations into an extension openpyxl drops on load, so without the
    rebuild a re-priced workbook would come back with no dropdowns at all.
    Returns the number of printings re-priced."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = load_workbook(path)
    ws, opts = wb["Cards"], wb["Options"]
    old_opts_max = opts.max_row

    # data rows end at the Total row; ranges come from each row's Set formula
    data_rows, ranges = [], {}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a == "Total":
            break
        if not a or str(a).startswith("Prices:"):
            continue
        data_rows.append(r)
        m = OPT_RANGE_RE.search(str(ws.cell(r, 3).value or ""))
        if m:
            ranges[r] = (int(m.group(1)), int(m.group(2)))

    # which rows did the seller correct (or did the original export fail on)?
    to_fix = []
    for r in data_rows:
        name = str(ws.cell(r, 1).value)
        printing = ws.cell(r, 2).value
        orig = ws.cell(r, 12).value
        labels = ([opts.cell(i, 1).value
                   for i in range(ranges[r][0], ranges[r][1] + 1)]
                  if r in ranges else [])
        if (orig and name != orig) or (printing and printing not in labels) \
                or not labels:
            to_fix.append((r, name, printing))

    # re-resolve corrected rows: same lookup chain as a fresh scan
    fixed = []
    for r, name, printing in to_fix:
        m = SET_CODE_RE.search(str(printing or ""))
        code = m.group(0).upper() if m else None
        cname, cands = resolve_candidates(name, code)
        fixed.append({"row": r, "typed": str(printing or ""), "name": cname,
                      "sets": cands, "set_code": "", "rarity": "", "guess": ""})
    if fixed:
        build_options(fixed, progress=None)  # prices the new options
        for card in fixed:
            r = card["row"]
            if not card["options"]:
                continue
            lo = opts.max_row + 1
            for o in card["options"]:
                opts.append(_option_row(o))
            hi = opts.max_row
            ranges[r] = (lo, hi)
            ws.cell(r, 1).value = card["name"]
            ws.cell(r, 12).value = card["name"]
            labels = [o["label"] for o in card["options"]]
            typed = card["typed"].strip().lower()
            pick = next((l for l in labels if l.lower() == typed), None)
            if pick is None and len(labels) == 1:
                pick = labels[0]
            ws.cell(r, 2).value = pick
            if pick is None:
                ws.cell(r, 2).fill = PatternFill("solid", fgColor=TINT)
            if not ws.cell(r, 4).value:
                ws.cell(r, 4).value = DEFAULT_CONDITION

    # rebuild every dropdown + formula set (Excel's re-save strips them)
    ws.data_validations.dataValidation = []
    cond_dv = DataValidation(
        type="list", formula1='"' + ",".join(CONDITIONS) + '"',
        allow_blank=False, showErrorMessage=True)
    ws.add_data_validation(cond_dv)
    for r in data_rows:
        if r in ranges:
            _wire_row(ws, r, *ranges[r])
            cond_dv.add(ws.cell(r, 4).coordinate)

    # re-price the pre-existing option rows (new ones were priced just now)
    rows = []  # (options_row, product_id)
    for r in range(2, old_opts_max + 1):
        m = PRODUCT_URL_RE.search(opts.cell(r, 3).value or "")
        if m:
            rows.append((r, int(m.group(1))))
    prices = _price_jobs({(pid, cond) for _, pid in rows for cond in CONDITIONS},
                         progress)
    for r, pid in rows:
        for i, cond in enumerate(CONDITIONS):
            priced = prices.get((pid, cond))
            opts.cell(r, 4 + i).value = priced["price"] if priced else None
            opts.cell(r, 9 + i).value = (
                "" if priced is None else ("yes" if priced["verified"] else "no"))

    if timestamp and str(ws["A2"].value or "").startswith("Prices:"):
        ws["A2"] = ("Prices: lowest verified TCGPlayer listing in the chosen "
                    f"condition, {timestamp}. A blank price means no verified "
                    "listing right now.")
    wb.save(path)
    return len(rows) + sum(len(c["options"]) for c in fixed)


def selftest():
    import os
    import tempfile
    from openpyxl import load_workbook

    global _IDX
    _IDX = ({("MZMU-EN003", "ultra rare"):
             {"productId": 1, "url": "u", "name": "Stare of the Snake Hair", "group": "Set A"},
             ("SBLS-EN026", "common"):
             {"productId": 2, "url": "u", "name": "The Snake Hair", "group": "Set B"}}, {})
    try:
        cands = tcgp_candidates("Stare of the Snake-Hair", "")   # name-only mode
        assert cands == [{"set_code": "MZMU-EN003", "set_name": "Set A",
                          "rarities": ["Ultra Rare"],
                          "card_name": "Stare of the Snake Hair"}], cands
        assert tcgp_candidates("Stare of the Snake-Hair", "SBLS-EN026") == []  # code excludes
    finally:
        _IDX = None

    # merge mode: no code read + ygoprodeck list incomplete → TCGPlayer
    # printings union in (surviving TCGPlayer's own typos), near-name
    # neighbours stay out
    import scan_cards
    real_db = scan_cards.db_lookup
    scan_cards.db_lookup = lambda name, code, snippet=None: (
        name, [{"set_code": "SP14-EN018", "set_name": "Star Pack 2014",
                "rarities": ["Common"]}])
    _IDX = ({("DUAD-EN004", "ultra rare"):
             {"productId": 1, "url": "u", "name": "Dodododo Warrior",  # typo is TCGPlayer's
              "group": "Duelist's Advance"},
             ("MAGI-EN001", "common"):
             {"productId": 2, "url": "u", "name": "Dododo Warrior Girl",
              "group": "G"}}, {})
    try:
        _, cands = resolve_candidates("Dododo Warrior", None)
        codes = {c["set_code"] for c in cands}
        assert "SP14-EN018" in codes and "DUAD-EN004" in codes, codes
        assert "MAGI-EN001" not in codes, codes  # 'Girl' variant excluded at 0.9

        # a misread code that matches nothing degrades to the same name-merge
        _, cands = resolve_candidates("Dododo Warrior", "TN19-EN001")
        codes = {c["set_code"] for c in cands}
        assert "DUAD-EN004" in codes and "SP14-EN018" in codes, codes
    finally:
        scan_cards.db_lookup = real_db
        _IDX = None

    def opt(label, nm_price, pid=0):
        return {"label": label, "set_name": "25th Anniversary Rarity Collection",
                "url": f"https://www.tcgplayer.com/product/{pid}/x" if nm_price else None,
                "prices": {c: (nm_price if c == "Near Mint" else None) for c in CONDITIONS},
                "verified": {c: (True if c == "Near Mint" and nm_price else None)
                             for c in CONDITIONS}}

    cards = [
        {"name": "Fallen of Albaz", "photo": "a.jpg", "set_code": "RA01-EN021",
         "rarity": "Secret Rare", "sets": [],
         "options": [opt("RA01-EN021 · Secret Rare", 1.23, 524540),
                     opt("RA01-EN021 · Ultra Rare", None)]},
        {"name": "Mystery Card", "photo": "a.jpg", "set_code": "", "rarity": "",
         "sets": [], "options": []},
        {"name": "Witness of the Ancient", "photo": "b.jpg", "set_code": "CORI-EN012",
         "rarity": "Ultra Rare", "guess": "ultra rare (extended art)", "qty": 3,
         "sets": [],
         "options": [opt("CORI-EN012 · Ultra Rare", 0.50, 611001),
                     opt("CORI-EN012 · Ultra Rare (Extended Art)", 4.00, 611002)]},
    ]
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "t.xlsx")
        write_workbook(p, cards, "2026-07-19")
        wb = load_workbook(p)
        ws = wb["Cards"]
        assert wb["Options"].sheet_state == "hidden"
        assert wb["Options"]["D1"].value == "Near Mint"           # MATCH header row
        assert ws["B3"].value == "RA01-EN021 · Secret Rare"       # preselected
        assert ws["D3"].value == "Near Mint"                      # default condition
        assert ws["E3"].value == 1 and ws["E5"].value == 3        # qty column
        assert "IF(INDEX" in ws["F3"].value and "MATCH($D3" in ws["F3"].value, ws["F3"].value
        assert ws["G3"].value == '=IFERROR($E3*$F3,"")'           # qty x unit
        # one printing dropdown per card with options + the shared condition list
        assert len(ws.data_validations.dataValidation) == 3
        assert ws["B4"].value is None                             # unresolved left blank
        assert ws["B5"].value == "CORI-EN012 · Ultra Rare (Extended Art)"  # EA guess wins

        # refresh: seller edits survive, prices change, corrected/unresolved
        # rows re-resolve — all offline (lookups and pricing stubbed)
        import scan_cards
        from ygo_tcgplayer_pricer import tcgplayer_catalog
        ws["B3"] = "RA01-EN021 · Ultra Rare"  # dropdown correction (valid label)
        ws["D3"] = "Lightly Played"           # condition correction
        wb.save(p)
        real = (tcgplayer_pricing.get_lowest_price_safe, scan_cards.db_lookup,
                tcgplayer_catalog.find_group_id)
        tcgplayer_pricing.get_lowest_price_safe = (
            lambda pid, condition=None, **kw: {"price": 9.99, "seller": "s", "verified": True})
        # the Mystery Card row (no options) re-resolves via these stubs
        scan_cards.db_lookup = (
            lambda name, code, snippet=None:
            ("Found Card", [{"set_code": "NEW1-EN001", "set_name": "New Set",
                             "rarities": ["Rare"]}]) if name == "Mystery Card"
            else (name, []))
        tcgplayer_catalog.find_group_id = lambda s: None
        _IDX = ({("NEW1-EN001", "rare"):
                 {"productId": 9, "url": "https://www.tcgplayer.com/product/9/x",
                  "name": "Found Card", "group": "New Set"}}, {})
        try:
            n = refresh_workbook(p, "2026-08-01")
        finally:
            (tcgplayer_pricing.get_lowest_price_safe, scan_cards.db_lookup,
             tcgplayer_catalog.find_group_id) = real
            _IDX = None
        wb2 = load_workbook(p)
        assert n == 4, n                              # 3 URL rows + 1 new option
        assert wb2["Options"]["D2"].value == 9.99                 # re-priced
        assert wb2["Cards"]["B3"].value == "RA01-EN021 · Ultra Rare"   # seller choice kept
        assert wb2["Cards"]["D3"].value == "Lightly Played"
        assert wb2["Cards"]["A4"].value == "Found Card"           # name corrected
        assert wb2["Cards"]["B4"].value == "NEW1-EN001 · Rare"    # new dropdown pick
        assert "VLOOKUP" in str(wb2["Cards"]["C4"].value)         # row wired up
        # dropdowns rebuilt: 3 printing DVs + the shared condition DV
        assert len(wb2["Cards"].data_validations.dataValidation) == 4
        assert "2026-08-01" in wb2["Cards"]["A2"].value
        assert wb2["Cards"]["A7"].value == "Total"
        assert wb2["Cards"]["G7"].value == "=SUM(G3:G5)", wb2["Cards"]["G7"].value
        # offer block: editable inputs, class-split bulk, discounted remainder
        assert 'MATCH(TRIM(MID($B3' in wb2["Cards"]["K3"].value
        assert '"Starfoil Rare"' in wb2["Cards"]["K3"].value
        assert wb2["Cards"]["F9"].value == 1.0                    # pricepoint input
        assert wb2["Cards"]["F10"].value == 0.65                  # rate input
        assert '"cr"' in wb2["Cards"]["E11"].value
        assert wb2["Cards"]["G11"].value == "=E11*5/1000"
        assert wb2["Cards"]["G12"].value == "=E12*30/1000"
        assert wb2["Cards"]["A14"].value == "OFFER TOTAL"
        assert wb2["Cards"]["G14"].value == "=G11+G12+G13*$F$10"
        assert wb2["Cards"].column_dimensions["K"].hidden
    print("selftest OK")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--selftest":
        selftest()
